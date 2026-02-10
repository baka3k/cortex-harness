#!/usr/bin/env python3
"""
Skeleton pipeline for Excel table detection + header normalization.
Focus: accuracy-first parsing for numeric/date values, merged cells,
and multi-header tables within one sheet.

Usage:
  python extractor/excel/excel_table_pipeline_skeleton.py --xlsx /path/to/file.xlsx
  python extractor/excel/excel_table_pipeline_skeleton.py --xlsx /path/to/file.xlsx --sheet "Sheet1"
  python extractor/excel/excel_table_pipeline_skeleton.py --xlsx /path/to/file.xlsx --key-columns "ID,Project Code"
  python extractor/excel/excel_table_pipeline_skeleton.py --xlsx /path/to/file.xlsx --max-row-gap 1 --max-col-gap 2
  python extractor/excel/excel_table_pipeline_skeleton.py --xlsx /path/to/file.xlsx --skip-footer-rows
  python extractor/excel/excel_table_pipeline_skeleton.py --xlsx /path/to/file.xlsx --entity-provider gliner
  python extractor/excel/excel_table_pipeline_skeleton.py --xlsx /path/to/file.xlsx --entity-column-map "Department:ORG,Owner:PERSON"
"""

from __future__ import annotations

import argparse
import hashlib
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook


@dataclass
class CellPayload:
    raw: Any
    formatted: str
    formula: Optional[str]
    data_type: Optional[str]
    is_date: bool
    number_format: Optional[str]


@dataclass
class TableRegion:
    table_id: str
    min_row: int
    max_row: int
    min_col: int
    max_col: int


@dataclass
class RowRecord:
    row_id: str
    row_hash: str
    values: Dict[str, str]
    raw_values: Dict[str, Any]
    cell_meta: Dict[str, Dict[str, Any]]
    serialized: str
    row_index: int
    is_footer: bool


def _normalize_whitespace(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def _format_number(value: Any) -> str:
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(Decimal(str(value)))
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def _format_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        return _format_number(value)
    return _normalize_whitespace(str(value))


def _is_non_empty(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str) and not value.strip():
        return False
    return True


def _is_text_like(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    return bool(re.search(r"[A-Za-z]", value))


def _is_numeric_like(value: Any) -> bool:
    return isinstance(value, (int, float, Decimal)) and not isinstance(value, bool)


def load_workbooks(xlsx_path: Path):
    # Accuracy-first: open formula workbook with read_only=False to access merged cells.
    wb_values = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    wb_formula = load_workbook(filename=str(xlsx_path), data_only=False, read_only=False)
    return wb_values, wb_formula


def build_cell_map(sheet_values, sheet_formula) -> Dict[Tuple[int, int], CellPayload]:
    cells: Dict[Tuple[int, int], CellPayload] = {}
    for row in sheet_values.iter_rows():
        for cell in row:
            formula_cell = sheet_formula.cell(row=cell.row, column=cell.column)
            raw = cell.value
            formula = formula_cell.value if formula_cell.data_type == "f" else None
            formatted = _format_value(raw)
            payload = CellPayload(
                raw=raw,
                formatted=formatted,
                formula=formula,
                data_type=cell.data_type,
                is_date=cell.is_date,
                number_format=getattr(formula_cell, "number_format", None),
            )
            cells[(cell.row, cell.column)] = payload
    return cells


def extract_merged_ranges(sheet_formula) -> List[Tuple[int, int, int, int]]:
    ranges: List[Tuple[int, int, int, int]] = []
    for merged in sheet_formula.merged_cells.ranges:
        ranges.append((merged.min_row, merged.max_row, merged.min_col, merged.max_col))
    return ranges


def apply_merged_cell_fill(
    cells: Dict[Tuple[int, int], CellPayload],
    merged_ranges: List[Tuple[int, int, int, int]],
) -> None:
    for min_row, max_row, min_col, max_col in merged_ranges:
        anchor = cells.get((min_row, min_col))
        if not anchor or not _is_non_empty(anchor.raw):
            continue
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                payload = cells.get((r, c))
                if payload and _is_non_empty(payload.raw):
                    continue
                cells[(r, c)] = anchor


def detect_tables(
    cells: Dict[Tuple[int, int], CellPayload],
    sheet_name: str,
    min_rows: int = 2,
    min_cols: int = 2,
    max_row_gap: int = 1,
    max_col_gap: int = 1,
) -> List[TableRegion]:
    coords = {
        (r, c)
        for (r, c), payload in cells.items()
        if _is_non_empty(payload.raw)
    }
    components: List[List[Tuple[int, int]]] = []
    table_index = 0
    while coords:
        start = coords.pop()
        stack = [start]
        component = [start]
        while stack:
            r, c = stack.pop()
            for nr, nc in ((r - 1, c), (r + 1, c), (r, c - 1), (r, c + 1)):
                if (nr, nc) in coords:
                    coords.remove((nr, nc))
                    stack.append((nr, nc))
                    component.append((nr, nc))
        components.append(component)

    boxes = []
    for comp in components:
        rows = [r for r, _ in comp]
        cols = [c for _, c in comp]
        boxes.append((min(rows), max(rows), min(cols), max(cols)))

    merged = _merge_boxes_by_gap(boxes, max_row_gap=max_row_gap, max_col_gap=max_col_gap)

    tables: List[TableRegion] = []
    for min_row, max_row, min_col, max_col in merged:
        if (max_row - min_row + 1) < min_rows or (max_col - min_col + 1) < min_cols:
            continue
        table_index += 1
        table_id = f"{sheet_name}::table_{table_index}::{min_row}:{max_row}:{min_col}:{max_col}"
        tables.append(
            TableRegion(
                table_id=table_id,
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            )
        )
    return tables


def _merge_boxes_by_gap(
    boxes: List[Tuple[int, int, int, int]],
    max_row_gap: int,
    max_col_gap: int,
) -> List[Tuple[int, int, int, int]]:
    if not boxes:
        return []
    changed = True
    boxes = boxes[:]
    while changed:
        changed = False
        new_boxes: List[Tuple[int, int, int, int]] = []
        used = [False] * len(boxes)
        for i, box in enumerate(boxes):
            if used[i]:
                continue
            min_r, max_r, min_c, max_c = box
            for j in range(i + 1, len(boxes)):
                if used[j]:
                    continue
                min_r2, max_r2, min_c2, max_c2 = boxes[j]
                if _boxes_close(
                    (min_r, max_r, min_c, max_c),
                    (min_r2, max_r2, min_c2, max_c2),
                    max_row_gap=max_row_gap,
                    max_col_gap=max_col_gap,
                ):
                    min_r = min(min_r, min_r2)
                    max_r = max(max_r, max_r2)
                    min_c = min(min_c, min_c2)
                    max_c = max(max_c, max_c2)
                    used[j] = True
                    changed = True
            used[i] = True
            new_boxes.append((min_r, max_r, min_c, max_c))
        boxes = new_boxes
    return boxes


def _boxes_close(
    a: Tuple[int, int, int, int],
    b: Tuple[int, int, int, int],
    max_row_gap: int,
    max_col_gap: int,
) -> bool:
    min_r1, max_r1, min_c1, max_c1 = a
    min_r2, max_r2, min_c2, max_c2 = b
    row_gap = max(0, max(min_r2 - max_r1, min_r1 - max_r2) - 1)
    col_gap = max(0, max(min_c2 - max_c1, min_c1 - max_c2) - 1)
    return row_gap <= max_row_gap and col_gap <= max_col_gap


def extract_table_matrix(
    cells: Dict[Tuple[int, int], CellPayload],
    table: TableRegion,
) -> List[List[CellPayload]]:
    matrix: List[List[CellPayload]] = []
    for r in range(table.min_row, table.max_row + 1):
        row: List[CellPayload] = []
        for c in range(table.min_col, table.max_col + 1):
            row.append(cells.get((r, c), CellPayload(None, "", None, None, False, None)))
        matrix.append(row)
    return matrix


def detect_header_rows(
    table_rows: List[List[CellPayload]],
    max_header_rows: int = 3,
    min_text_ratio: float = 0.5,
    max_numeric_ratio: float = 0.5,
) -> List[int]:
    header_rows: List[int] = []
    for idx, row in enumerate(table_rows[:max_header_rows]):
        non_empty = [cell for cell in row if _is_non_empty(cell.raw)]
        if not non_empty:
            continue
        if len(non_empty) <= 1:
            # Likely title row, skip it and keep searching.
            continue
        text_count = sum(1 for cell in non_empty if _is_text_like(cell.formatted))
        numeric_count = sum(1 for cell in non_empty if _is_numeric_like(cell.raw))
        text_ratio = text_count / len(non_empty)
        numeric_ratio = numeric_count / len(non_empty)
        if text_ratio >= min_text_ratio and numeric_ratio <= max_numeric_ratio:
            header_rows.append(idx)
        else:
            break
    if not header_rows:
        header_rows = [0]
    return header_rows


def normalize_headers(
    table_rows: List[List[CellPayload]],
    header_rows: Sequence[int],
) -> List[str]:
    if not table_rows:
        return []
    col_count = len(table_rows[0])
    header_matrix = _build_header_matrix(table_rows, header_rows, col_count)
    headers: List[str] = []
    for col_idx in range(col_count):
        parts: List[str] = []
        for row_labels in header_matrix:
            label = row_labels[col_idx]
            if label:
                parts.append(label)
        header = " / ".join(parts).strip() if parts else f"col_{col_idx + 1}"
        headers.append(header)
    return _make_unique_headers(headers)


def _build_header_matrix(
    table_rows: List[List[CellPayload]],
    header_rows: Sequence[int],
    col_count: int,
) -> List[List[str]]:
    matrix: List[List[str]] = []
    for row_idx in header_rows:
        labels: List[str] = []
        last_label = ""
        for col_idx in range(col_count):
            cell = table_rows[row_idx][col_idx]
            label = _normalize_whitespace(cell.formatted)
            if not label and last_label:
                label = last_label
            if label:
                last_label = label
            labels.append(label)
        matrix.append(labels)
    return matrix


def _make_unique_headers(headers: List[str]) -> List[str]:
    seen: Dict[str, int] = {}
    unique: List[str] = []
    for name in headers:
        key = name or "col"
        count = seen.get(key, 0)
        seen[key] = count + 1
        if count == 0:
            unique.append(key)
        else:
            unique.append(f"{key}_{count + 1}")
    return unique


def iter_data_rows(
    table_rows: List[List[CellPayload]],
    headers: Sequence[str],
    header_rows: Sequence[int],
    table_id: str,
    key_columns: Optional[Sequence[str]] = None,
    skip_footer_rows: bool = False,
) -> Iterable[RowRecord]:
    start_idx = max(header_rows) + 1 if header_rows else 1
    for row_offset, row in enumerate(table_rows[start_idx:], start=start_idx):
        values: Dict[str, str] = {}
        raw_values: Dict[str, Any] = {}
        cell_meta: Dict[str, Dict[str, Any]] = {}
        has_content = False
        for col_idx, header in enumerate(headers):
            cell = row[col_idx]
            if _is_non_empty(cell.raw):
                has_content = True
            values[header] = cell.formatted
            raw_values[header] = cell.raw
            cell_meta[header] = {
                "data_type": cell.data_type,
                "is_date": cell.is_date,
                "formula": cell.formula,
                "number_format": cell.number_format,
            }
        if not has_content:
            continue
        is_footer = _is_footer_row(values)
        if skip_footer_rows and is_footer:
            continue
        row_id, row_hash = compute_row_identity(
            values, table_id, key_columns=key_columns
        )
        serialized = serialize_row(values)
        yield RowRecord(
            row_id=row_id,
            row_hash=row_hash,
            values=values,
            raw_values=raw_values,
            cell_meta=cell_meta,
            serialized=serialized,
            row_index=row_offset,
            is_footer=is_footer,
        )


def compute_row_identity(
    values: Dict[str, str],
    table_id: str,
    key_columns: Optional[Sequence[str]] = None,
) -> Tuple[str, str]:
    key_columns = [c.strip() for c in (key_columns or []) if c.strip()]
    if key_columns:
        normalized = _normalize_key_columns(values)
        key_parts = [normalized.get(col.lower(), "") for col in key_columns]
        if any(part for part in key_parts):
            row_id = _stable_hash([table_id] + key_parts)
            row_hash = _stable_hash([table_id] + list(values.values()))
            return row_id, row_hash
    row_hash = _stable_hash([table_id] + list(values.values()))
    return row_hash, row_hash


def _stable_hash(parts: Sequence[str]) -> str:
    h = hashlib.sha1()
    for part in parts:
        h.update(part.encode("utf-8"))
        h.update(b"|")
    return h.hexdigest()


def _normalize_key_columns(values: Dict[str, str]) -> Dict[str, str]:
    normalized: Dict[str, str] = {}
    for key, value in values.items():
        normalized[key.lower()] = value
    return normalized


def serialize_row(values: Dict[str, str]) -> str:
    parts = []
    for key, value in values.items():
        if value:
            parts.append(f"{key}: {value}")
    return ". ".join(parts)


def _is_footer_row(values: Dict[str, str]) -> bool:
    footer_tokens = ("total", "sum", "tong", "subtotal", "grand total")
    for value in values.values():
        if isinstance(value, str) and value:
            lower = value.lower()
            if any(token in lower for token in footer_tokens):
                return True
    return False


def parse_entity_columns(raw: str) -> List[str]:
    if not raw:
        return []
    return [part.strip() for part in raw.split(",") if part.strip()]


def parse_entity_column_map(raw: str) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    if not raw:
        return mapping
    items = [item.strip() for item in raw.split(",") if item.strip()]
    for item in items:
        if ":" in item:
            col, etype = item.split(":", 1)
        elif "=" in item:
            col, etype = item.split("=", 1)
        else:
            col, etype = item, "COLUMN"
        col = col.strip()
        etype = etype.strip() or "COLUMN"
        if col:
            mapping[col] = etype
    return mapping


def build_column_entities(
    values: Dict[str, str],
    column_map: Dict[str, str],
    column_list: Sequence[str],
    default_type: str,
) -> List[Dict[str, str]]:
    entities: List[Dict[str, str]] = []
    lower_values = {k.lower(): v for k, v in values.items()}
    for col, etype in column_map.items():
        value = lower_values.get(col.lower(), "")
        if value:
            entities.append({"name": value, "type": etype})
    for col in column_list:
        value = lower_values.get(col.lower(), "")
        if value:
            entities.append({"name": value, "type": default_type})
    return entities


def _dedup_entities(entities: List[Dict[str, str]]) -> List[Dict[str, str]]:
    seen = set()
    unique: List[Dict[str, str]] = []
    for ent in entities:
        name = str(ent.get("name", "")).strip()
        etype = str(ent.get("type", "")).strip()
        if not name:
            continue
        key = (etype.lower(), name.lower())
        if key in seen:
            continue
        seen.add(key)
        unique.append({"name": name, "type": etype or "UNKNOWN"})
    return unique


def extract_entities_for_rows(
    rows: List[RowRecord],
    provider: str,
    batch_size: int,
    gliner_labels: Optional[str],
    gliner_threshold: float,
    gliner_model: Optional[str],
    spacy_model: Optional[str],
    spacy_ruler: Optional[str],
    column_map: Dict[str, str],
    column_list: Sequence[str],
    column_default_type: str,
) -> List[List[Dict[str, str]]]:
    base_entities = [
        build_column_entities(row.values, column_map, column_list, column_default_type)
        for row in rows
    ]

    if provider == "none":
        return [_dedup_entities(items) for items in base_entities]

    texts = [row.serialized for row in rows]
    extracted: List[List[Dict[str, str]]] = [[] for _ in rows]

    if provider == "gliner":
        from entity_extractors import (
            build_gliner_model,
            extract_entities_gliner_batch,
            parse_gliner_labels,
        )

        label_list = parse_gliner_labels(gliner_labels)
        model_instance = build_gliner_model(gliner_model or "urchade/gliner_large-v2.1")
        batch_size = max(1, batch_size)
        for start in range(0, len(texts), batch_size):
            batch_texts = texts[start : start + batch_size]
            batch_entities = extract_entities_gliner_batch(
                batch_texts,
                labels=label_list,
                threshold=gliner_threshold,
                gliner_model=model_instance,
            )
            for idx, entities in enumerate(batch_entities, start=start):
                extracted[idx] = entities
    elif provider == "spacy":
        from entity_extractors import extract_entities_spacy

        model = spacy_model or "en_core_web_sm"
        for idx, text in enumerate(texts):
            entities, _ = extract_entities_spacy(text, model=model, ruler_json=spacy_ruler)
            extracted[idx] = entities
    elif provider == "gemini":
        from entity_extractors import extract_entities_gemini

        for idx, text in enumerate(texts):
            entities, _ = extract_entities_gemini(text)
            extracted[idx] = entities
    elif provider == "langextract":
        from entity_extractors import extract_entities_langextract

        for idx, text in enumerate(texts):
            entities, _ = extract_entities_langextract(text)
            extracted[idx] = entities
    else:
        raise ValueError(f"Unknown entity provider: {provider}")

    merged: List[List[Dict[str, str]]] = []
    for base, extra in zip(base_entities, extracted, strict=True):
        merged.append(_dedup_entities(base + extra))
    return merged


def run_pipeline(
    xlsx_path: Path,
    sheet_name: Optional[str],
    key_columns: List[str],
    min_rows: int,
    min_cols: int,
    max_row_gap: int,
    max_col_gap: int,
    max_header_rows: int,
    min_text_ratio: float,
    max_numeric_ratio: float,
    skip_footer_rows: bool,
    entity_provider: str,
    entity_batch_size: int,
    entity_columns: List[str],
    entity_column_map: Dict[str, str],
    entity_column_type: str,
    gliner_labels: Optional[str],
    gliner_threshold: float,
    gliner_model: Optional[str],
    spacy_model: Optional[str],
    spacy_ruler: Optional[str],
):
    wb_values, wb_formula = load_workbooks(xlsx_path)
    sheets = (
        [wb_values[sheet_name]] if sheet_name else list(wb_values.worksheets)
    )
    for sheet in sheets:
        sheet_formula = wb_formula[sheet.title]
        cells = build_cell_map(sheet, sheet_formula)
        merged_ranges = extract_merged_ranges(sheet_formula)
        apply_merged_cell_fill(cells, merged_ranges)
        tables = detect_tables(
            cells,
            sheet.title,
            min_rows=min_rows,
            min_cols=min_cols,
            max_row_gap=max_row_gap,
            max_col_gap=max_col_gap,
        )
        print(f"Sheet: {sheet.title} -> tables: {len(tables)}")
        for table in tables:
            matrix = extract_table_matrix(cells, table)
            header_rows = detect_header_rows(
                matrix,
                max_header_rows=max_header_rows,
                min_text_ratio=min_text_ratio,
                max_numeric_ratio=max_numeric_ratio,
            )
            headers = normalize_headers(matrix, header_rows)
            rows = list(
                iter_data_rows(
                    matrix,
                    headers,
                    header_rows,
                    table.table_id,
                    key_columns=key_columns,
                    skip_footer_rows=skip_footer_rows,
                )
            )
            row_count = len(rows)
            entity_count = 0
            if rows and (entity_provider != "none" or entity_columns or entity_column_map):
                entities_by_row = extract_entities_for_rows(
                    rows,
                    provider=entity_provider,
                    batch_size=entity_batch_size,
                    gliner_labels=gliner_labels,
                    gliner_threshold=gliner_threshold,
                    gliner_model=gliner_model,
                    spacy_model=spacy_model,
                    spacy_ruler=spacy_ruler,
                    column_map=entity_column_map,
                    column_list=entity_columns,
                    column_default_type=entity_column_type,
                )
                entity_count = sum(len(items) for items in entities_by_row)
            print(
                f"  {table.table_id} rows={row_count} "
                f"header_rows={header_rows} cols={len(headers)} entities={entity_count}"
            )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True, help="Path to .xlsx file")
    parser.add_argument("--sheet", help="Sheet name to parse (optional)")
    parser.add_argument(
        "--key-columns",
        help="Comma-separated list of key columns for stable row_id",
        default="",
    )
    parser.add_argument("--min-rows", type=int, default=2, help="Min rows per table")
    parser.add_argument("--min-cols", type=int, default=2, help="Min cols per table")
    parser.add_argument(
        "--max-row-gap",
        type=int,
        default=1,
        help="Allowed empty row gap when merging table blocks",
    )
    parser.add_argument(
        "--max-col-gap",
        type=int,
        default=1,
        help="Allowed empty col gap when merging table blocks",
    )
    parser.add_argument(
        "--max-header-rows",
        type=int,
        default=3,
        help="Max header rows to consider",
    )
    parser.add_argument(
        "--min-text-ratio",
        type=float,
        default=0.5,
        help="Min text-like ratio for header detection",
    )
    parser.add_argument(
        "--max-numeric-ratio",
        type=float,
        default=0.5,
        help="Max numeric ratio for header detection",
    )
    parser.add_argument(
        "--skip-footer-rows",
        action="store_true",
        help="Skip rows that look like totals/subtotals",
    )
    parser.add_argument(
        "--entity-provider",
        choices=["none", "gliner", "spacy", "gemini", "langextract"],
        default="none",
        help="Entity extraction provider",
    )
    parser.add_argument(
        "--entity-batch-size",
        type=int,
        default=16,
        help="Batch size for entity extraction (gliner only)",
    )
    parser.add_argument(
        "--entity-columns",
        default="",
        help="Comma-separated column names to treat as entities (default type COLUMN)",
    )
    parser.add_argument(
        "--entity-column-map",
        default="",
        help="Comma-separated column:type mapping, e.g. Department:ORG,Owner:PERSON",
    )
    parser.add_argument(
        "--entity-column-type",
        default="COLUMN",
        help="Default entity type for --entity-columns",
    )
    parser.add_argument(
        "--gliner-labels",
        default=None,
        help="Comma-separated labels or path to labels file",
    )
    parser.add_argument(
        "--gliner-threshold",
        type=float,
        default=0.3,
        help="GLiNER threshold",
    )
    parser.add_argument(
        "--gliner-model",
        default=None,
        help="GLiNER model name or local path",
    )
    parser.add_argument(
        "--spacy-model",
        default=None,
        help="spaCy model name (default: en_core_web_sm)",
    )
    parser.add_argument(
        "--spacy-ruler",
        default=None,
        help="spaCy ruler JSON file/dir (optional)",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)
    key_columns = [c.strip() for c in args.key_columns.split(",") if c.strip()]
    entity_columns = parse_entity_columns(args.entity_columns)
    entity_column_map = parse_entity_column_map(args.entity_column_map)
    run_pipeline(
        xlsx_path,
        args.sheet,
        key_columns,
        min_rows=args.min_rows,
        min_cols=args.min_cols,
        max_row_gap=args.max_row_gap,
        max_col_gap=args.max_col_gap,
        max_header_rows=args.max_header_rows,
        min_text_ratio=args.min_text_ratio,
        max_numeric_ratio=args.max_numeric_ratio,
        skip_footer_rows=args.skip_footer_rows,
        entity_provider=args.entity_provider,
        entity_batch_size=args.entity_batch_size,
        entity_columns=entity_columns,
        entity_column_map=entity_column_map,
        entity_column_type=args.entity_column_type,
        gliner_labels=args.gliner_labels,
        gliner_threshold=args.gliner_threshold,
        gliner_model=args.gliner_model,
        spacy_model=args.spacy_model,
        spacy_ruler=args.spacy_ruler,
    )


if __name__ == "__main__":
    main()
