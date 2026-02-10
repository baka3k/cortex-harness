#!/usr/bin/env python3
"""
Skeleton pipeline for Excel table detection + header normalization.
Focus: accuracy-first parsing for numeric/date values.

Usage:
  python excel_table_pipeline_skeleton.py --xlsx /path/to/file.xlsx
  python excel_table_pipeline_skeleton.py --xlsx /path/to/file.xlsx --sheet "Sheet1"
  python excel_table_pipeline_skeleton.py --xlsx /path/to/file.xlsx --key-columns "ID,Project Code"
"""

from __future__ import annotations

import argparse
import hashlib
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook


@dataclass
class CellPayload:
    raw: Any
    formatted: str
    formula: Optional[str]
    data_type: Optional[str]
    is_date: bool


@dataclass
class TableRegion:
    table_id: str
    min_row: int
    max_row: int
    min_col: int
    max_col: int


@dataclass
class RowRecord:
    row_id: str
    row_hash: str
    values: Dict[str, str]
    raw_values: Dict[str, Any]
    row_index: int


def _normalize_whitespace(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def _format_number(value: Any) -> str:
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(Decimal(str(value)))
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def _format_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        return _format_number(value)
    return _normalize_whitespace(str(value))


def _is_non_empty(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str) and not value.strip():
        return False
    return True


def _is_text_like(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    return bool(re.search(r"[A-Za-z]", value))


def _is_numeric_like(value: Any) -> bool:
    return isinstance(value, (int, float, Decimal)) and not isinstance(value, bool)


def load_workbooks(xlsx_path: Path):
    wb_values = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    wb_formula = load_workbook(filename=str(xlsx_path), data_only=False, read_only=True)
    return wb_values, wb_formula


def build_cell_map(sheet_values, sheet_formula) -> Dict[Tuple[int, int], CellPayload]:
    cells: Dict[Tuple[int, int], CellPayload] = {}
    for row in sheet_values.iter_rows():
        for cell in row:
            formula_cell = sheet_formula.cell(row=cell.row, column=cell.column)
            raw = cell.value
            formula = formula_cell.value if formula_cell.data_type == "f" else None
            formatted = _format_value(raw)
            payload = CellPayload(
                raw=raw,
                formatted=formatted,
                formula=formula,
                data_type=cell.data_type,
                is_date=cell.is_date,
            )
            cells[(cell.row, cell.column)] = payload
    return cells


def detect_tables(
    cells: Dict[Tuple[int, int], CellPayload],
    sheet_name: str,
    min_rows: int = 2,
    min_cols: int = 2,
) -> List[TableRegion]:
    coords = {
        (r, c)
        for (r, c), payload in cells.items()
        if _is_non_empty(payload.raw)
    }
    tables: List[TableRegion] = []
    table_index = 0
    while coords:
        start = coords.pop()
        stack = [start]
        component = [start]
        while stack:
            r, c = stack.pop()
            for nr, nc in ((r - 1, c), (r + 1, c), (r, c - 1), (r, c + 1)):
                if (nr, nc) in coords:
                    coords.remove((nr, nc))
                    stack.append((nr, nc))
                    component.append((nr, nc))
        rows = [r for r, _ in component]
        cols = [c for _, c in component]
        min_row, max_row = min(rows), max(rows)
        min_col, max_col = min(cols), max(cols)
        if (max_row - min_row + 1) < min_rows or (max_col - min_col + 1) < min_cols:
            continue
        table_index += 1
        table_id = f"{sheet_name}::table_{table_index}::{min_row}:{max_row}:{min_col}:{max_col}"
        tables.append(
            TableRegion(
                table_id=table_id,
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            )
        )
    return tables


def extract_table_matrix(
    cells: Dict[Tuple[int, int], CellPayload],
    table: TableRegion,
) -> List[List[CellPayload]]:
    matrix: List[List[CellPayload]] = []
    for r in range(table.min_row, table.max_row + 1):
        row: List[CellPayload] = []
        for c in range(table.min_col, table.max_col + 1):
            row.append(cells.get((r, c), CellPayload(None, "", None, None, False)))
        matrix.append(row)
    return matrix


def detect_header_rows(
    table_rows: List[List[CellPayload]],
    max_header_rows: int = 3,
    min_text_ratio: float = 0.6,
    max_numeric_ratio: float = 0.4,
) -> List[int]:
    header_rows: List[int] = []
    for idx, row in enumerate(table_rows[:max_header_rows]):
        non_empty = [cell for cell in row if _is_non_empty(cell.raw)]
        if not non_empty:
            continue
        text_count = sum(1 for cell in non_empty if _is_text_like(cell.formatted))
        numeric_count = sum(1 for cell in non_empty if _is_numeric_like(cell.raw))
        text_ratio = text_count / len(non_empty)
        numeric_ratio = numeric_count / len(non_empty)
        if text_ratio >= min_text_ratio and numeric_ratio <= max_numeric_ratio:
            header_rows.append(idx)
        else:
            break
    if not header_rows:
        header_rows = [0]
    return header_rows


def normalize_headers(
    table_rows: List[List[CellPayload]],
    header_rows: Sequence[int],
) -> List[str]:
    if not table_rows:
        return []
    col_count = len(table_rows[0])
    headers: List[str] = []
    for col_idx in range(col_count):
        parts: List[str] = []
        for row_idx in header_rows:
            cell = table_rows[row_idx][col_idx]
            label = _normalize_whitespace(cell.formatted)
            if label:
                parts.append(label)
        header = " / ".join(parts).strip() if parts else f"col_{col_idx + 1}"
        headers.append(header)
    return _make_unique_headers(headers)


def _make_unique_headers(headers: List[str]) -> List[str]:
    seen: Dict[str, int] = {}
    unique: List[str] = []
    for name in headers:
        key = name or "col"
        count = seen.get(key, 0)
        seen[key] = count + 1
        if count == 0:
            unique.append(key)
        else:
            unique.append(f"{key}_{count + 1}")
    return unique


def iter_data_rows(
    table_rows: List[List[CellPayload]],
    headers: Sequence[str],
    header_rows: Sequence[int],
    table_id: str,
    key_columns: Optional[Sequence[str]] = None,
) -> Iterable[RowRecord]:
    start_idx = max(header_rows) + 1 if header_rows else 1
    for row_offset, row in enumerate(table_rows[start_idx:], start=start_idx):
        values: Dict[str, str] = {}
        raw_values: Dict[str, Any] = {}
        has_content = False
        for col_idx, header in enumerate(headers):
            cell = row[col_idx]
            if _is_non_empty(cell.raw):
                has_content = True
            values[header] = cell.formatted
            raw_values[header] = cell.raw
        if not has_content:
            continue
        row_id, row_hash = compute_row_identity(
            values, table_id, key_columns=key_columns
        )
        yield RowRecord(
            row_id=row_id,
            row_hash=row_hash,
            values=values,
            raw_values=raw_values,
            row_index=row_offset,
        )


def compute_row_identity(
    values: Dict[str, str],
    table_id: str,
    key_columns: Optional[Sequence[str]] = None,
) -> Tuple[str, str]:
    key_columns = [c.strip() for c in (key_columns or []) if c.strip()]
    if key_columns:
        key_parts = [values.get(col, "") for col in key_columns]
        if any(part for part in key_parts):
            row_id = _stable_hash([table_id] + key_parts)
            row_hash = _stable_hash([table_id] + list(values.values()))
            return row_id, row_hash
    row_hash = _stable_hash([table_id] + list(values.values()))
    return row_hash, row_hash


def _stable_hash(parts: Sequence[str]) -> str:
    h = hashlib.sha1()
    for part in parts:
        h.update(part.encode("utf-8"))
        h.update(b"|")
    return h.hexdigest()


def run_pipeline(xlsx_path: Path, sheet_name: Optional[str], key_columns: List[str]):
    wb_values, wb_formula = load_workbooks(xlsx_path)
    sheets = (
        [wb_values[sheet_name]] if sheet_name else list(wb_values.worksheets)
    )
    for sheet in sheets:
        sheet_formula = wb_formula[sheet.title]
        cells = build_cell_map(sheet, sheet_formula)
        tables = detect_tables(cells, sheet.title)
        print(f"Sheet: {sheet.title} -> tables: {len(tables)}")
        for table in tables:
            matrix = extract_table_matrix(cells, table)
            header_rows = detect_header_rows(matrix)
            headers = normalize_headers(matrix, header_rows)
            row_count = 0
            for _ in iter_data_rows(
                matrix,
                headers,
                header_rows,
                table.table_id,
                key_columns=key_columns,
            ):
                row_count += 1
            print(
                f"  {table.table_id} rows={row_count} "
                f"header_rows={header_rows} cols={len(headers)}"
            )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True, help="Path to .xlsx file")
    parser.add_argument("--sheet", help="Sheet name to parse (optional)")
    parser.add_argument(
        "--key-columns",
        help="Comma-separated list of key columns for stable row_id",
        default="",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)
    key_columns = [c.strip() for c in args.key_columns.split(",") if c.strip()]
    run_pipeline(xlsx_path, args.sheet, key_columns)


if __name__ == "__main__":
    main()
